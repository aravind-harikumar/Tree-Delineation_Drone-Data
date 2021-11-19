"""
..   module  :: Main Function
    :platform: Unix, Windows
    :synopsis: UAV optical data processing and analysis workflow    
..  :author  : Aravind Harikumar <aravind.harikumar@utoronto.ca>
"""
import io, sys, os
from openpyxl import load_workbook
from MetashapeProcessor import generateOrthophoto
from ProjectConstants import GlobalConstants as gc
# sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from ImageProcessor.RasterOperators import ReadImage
from ITCDetector import DataPreprocessor, \
        IndividualTreeDetection as ITC, \
        ITCAnalyze, \
        IndividualTreeDelineation as ITD, \
        IndividualTreeBuffering as ITB, \
        IndividualTreeSpanning as ITSpan, \
        IndividualTreeDelineationAll as ITD1
    
Dates = ['20170809','20171011','20180516','20180710','20181015'] # ['20170626','20170809','20171011','20180516','20180710','20181015']
Base_RawData_Folder = '/mnt/4TBHDD/StCasimir-Multispectral-Analysis/1_InputForAnalayis/'
Base_Processed_Data_MS = '/mnt/4TBHDD/StCasimir-Multispectral-Analysis/2_Analysis-And-Results/ProcessedMSData1/'
Base_Processed_Data_HS = '/mnt/4TBHDD/StCasimir-Multispectral-Analysis/2_Analysis-And-Results/ProcessedHSData/'

# Global parameters to control section of the code to be executed
runpart,runpart_analysis,skipCoregStep = False,True,True

'''#################################################################'''
'''           Generate Othorectfied Image and DEM                   '''
'''#################################################################'''
if(runpart):
    wb = load_workbook(filename = 'StCasmir_Data.xlsx')
    ws = wb['Sheet1']

    for index, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row)):
        # skip header row
        if(str(row[0].value).strip() not in Dates):
            continue
        else:
            print("Processing :" + str(row[0].value))
        # # set image details
        DataInfo = {
        'ID'                 : str(row[0].value), # User Defined Project Name
        'DataPath'           : str(row[2].value), # Root Data Folder
        'OutFolder'          : Base_RawData_Folder, # Output Folder
        'LoadContolPointFile': bool(str(row[4].value)),# Load Control Points (True/False)
        'ContolPointFile'    : '/mnt/4TBHDD/StCasimir-Multispectral-Analysis/0_SpruceUp-Raw-Data/GPS/only_targets/SCA_GPS_final_edit_petra.csv', # used if LoadContolPointFile == True
        'OpenExistingProj'   : True, # required to run from an intermediate step mentioned below (True/False)
        'Primary_Channel'    : 4, # Primary Channel
        'EPSGcrs'            : 'EPSG::32619', # Reference System EPSG % 32619 4326
        }
        # Preprocess raw optical data and generate nDSM and Orthophoto
        DataObj = generateOrthophoto.GenerateOrthoImage(DataInfo)
        
'''#################################################################'''
'''                       Stack bands / Tile Image                  '''
'''#################################################################'''
if(runpart_analysis):
    for date in Dates:
        FileDataInfo = {
            'RSDataType'        : 'Mulispectral', # Hyperspectral or Mulispectral
            'BaseFolder'        :  Base_RawData_Folder,
            'ImageFolder'       :  date, # 20180514, 20180710, 20181015
            'OrthoPhoto'        :  date + '.tif',  # 20180514/20180514.tif , 20180710/20180710.tif , 20181015/20181015.tif
            'nDSM'              : 'nDSM.tif', # '/mnt/4TBHDD/Agisoft/nDSM_Z19_UTM.tif',
            'dsmresolution'     :  0.25,
            'SelectedBands'     : ['Band1', 'Band2', 'Band3', 'Band4', 'Band5'], # or ALL
            'OutFolderHyper'    :  Base_Processed_Data_HS,
            'OutFolderMulti'    :  Base_Processed_Data_MS,
            'RefSpanImage'      : '/mnt/4TBHDD/StCasimir-Multispectral-Analysis/1_InputForAnalayis/ReferenceFiles/nDSM_Z19_UTM.tif',#'/home/ensmingerlabgpu/Desktop/AgisoftProjects/20180514/20180514.tif',
            'RefnDSMImage'      : '/mnt/4TBHDD/StCasimir-Multispectral-Analysis/1_InputForAnalayis/20170810-REFERENCE/nDSM.tif',  #'/mnt/4TBHDD/Agisoft/20180514_ndsm.tif',
            'StudyAreaShp'      : '/mnt/4TBHDD/StCasimir-Multispectral-Analysis/0_SpruceUp-Raw-Data/Site_info/ArcGIS/StCasimir_SpruceUp_tempRect1.shp', #
            'CoBaseFolder'      :  Base_Processed_Data_MS,
            'RefDateFolder'     :  Dates[0],
            'DataType'          :  'uint16',
            'SkipCorStep'       :  skipCoregStep,
            'SaveGCPFromAgisoft':  True,
            'SaveAgisoftFile'   :  False
        }
        print('Preprocessing - '  + FileDataInfo['ImageFolder'])
        DataObj = DataPreprocessor.Preprocessor(FileDataInfo)
        DataObj.PreProcessData()

    '''#################################################################'''
    '''              Coregister Data From All Aquisitions    (optional) '''
    '''#################################################################'''    
    # Coregister Data Aquisitions From All Dates
    ITCDelinObj = ITD1.ITCDelineationUtils(FileDataInfo) # names of dsm and orthphoto should be same , only date  different in name
    if(not(skipCoregStep)):
        ITCDelinObj.Coregister()
        ITCDelinObj.Coregister_orig_dim_dem()

    '''###################################################################'''
    '''           Detect Individual Tree Crowns From Optical Data         '''
    '''###################################################################'''
    for date in Dates:
        FileDataInfo = {
            'BaseFolder'            :  Base_Processed_Data_MS,
            'ImageDate'             :  date, # 20180514, 20180710, 20181015,
            'clipShpFile'           :  '/mnt/4TBHDD/StCasimir-Multispectral-Analysis/0_SpruceUp-Raw-Data/Site_info/ArcGIS/StCasimir_SpruceUp_temp_plots_1.shp',
            'ITCBufferSize'         :  1.5,
            'TreeTopInclusionBuffer':  2, # includes multiple bright points around N m of tree top
            'InterTreeSepeartion'   :  1, # removes redundant points within Nm distance
            'MinTreeHeight'         :  2, # (m)
            'FilterVariance'        :  1
        }
        print('ITC Detecting -' + FileDataInfo['ImageDate'])
        ITCObj = ITC.ITCUtils(FileDataInfo)
        ITCObj.DetectIndividaulTrees()

    '''#########################################################################################'''
    '''                                 Get fuzzy maps of crown                                 '''
    '''#########################################################################################'''
    for date in Dates:
        FileDataInfo = {
            'BaseFolder'        :  Base_Processed_Data_MS,
            'Date'              :  date, # 20180514, 20180710, 20181015
            'ClipSize'          :  4,
            'RefCenter'         :  [1480, 1991, 1407, 919, 11280], # average spectral signature of tree class
            'NumOfClusters'     :  2,
            'SkipCoregStep'     :  skipCoregStep,
            'R_ArgFilename'     : '/home/ensmingerlabgpu/Documents/Code_Fuzzy/RProgram/FCMMRFPYTHON.RFilename2.txt',
            'R_ScriptName'      : '/home/ensmingerlabgpu/Documents/Code_Fuzzy/RProgram/FCMMRFPYTHON2.R'
        }
        print('ITC Delineating -' + FileDataInfo['Date'])
        ITCObj = ITB.ITCBufferUtils(FileDataInfo)
        # Genereate fuzzy maps for the trees
        ITCObj.BufferIndividaulTrees() # run Simple FCM Implemted In a Python, 
        # ITCObj.RBufferIndividaulTrees() # or # run MRF-FCM implemented in a R Script

    '''#########################################################################################'''
    '''                      Delineate Individual Tree Crowns From FCM Output                 '''
    '''#########################################################################################'''
    for date in Dates:
        FileDataInfo = {
            'BaseFolder'        :  Base_Processed_Data_MS,
            'Date'              :  date,
            'RefShp'            : '/home/ensmingerlabgpu/Documents/MATLAB/BasicSnake_version2f/myfile_1.shp',
            'MatFilePath'       : '/home/ensmingerlabgpu/Documents/MATLAB/BasicSnake_version2f',
            'SkipCoregStep'     :  skipCoregStep,
            'IndexOperatorType' : 'AllPixelMean', # [AllPixelMean, or MaxNPixelMean]
            'Skip'              :  False # Skip formating of shape files 
        }
        print('ITC Delineating -' + FileDataInfo['Date'])
        ITCObj = ITSpan.ITCCrownShapeUtils(FileDataInfo)
        # Run individual crown delineation using spectral,spatial and temporal information
        ITCObj.GetCrownFromMatlab()
        # Generate remote sensing indice maps and format and organize the shape file
        ITCObj.GetCrownBuffer()

'''#################################################################'''
'''                          Analysis                               '''
'''#################################################################'''
FileDataInfo = {
    'BaseFolder'        :  Base_Processed_Data_MS,
    'Dates'             : ['20170626', '20170809','20171011','20180516','20180710','20181015'], # Dates = ['20180516','20180710','20181015'] 
    'MaxNeighbourDist'  : 1,
    'ref_file_shp'      : "/mnt/4TBHDD/StCasimir-Multispectral-Analysis/0_SpruceUp-Raw-Data/Site_info/ReferenceData/SC_TREE_REFERENCE_DATA_1.shp", #"/mnt/4TBHDD/Spruce_Up_New_Copy/SpruceUp/GPS/trees_&_targets/trees_ABD.shp"
    'ref_100_file_shp'  : "/mnt/4TBHDD/StCasimir-Multispectral-Analysis/0_SpruceUp-Raw-Data/Site_info/ReferenceData/SC_100_TREE_REFERENCE_DATA.shp",
    'ALL_REF_DATA'      : "/mnt/4TBHDD/100_Tree_Ref_Data/refined/Reference_Data_StCasimir_2017_2018.xlsx",
    'REF_SHEET_ID'      : "Pigment_Data", #  ['WaterPotential_Data', 'Pigment_Data', 'GasExchange_Fluorescence_Data'],
    'AnalysisYears'     : [2017,2018], # Yearly data to be considered for analysis
    'AnalysisMonths'    : ['June','August','October'], # Monthly data to be considered for analysis
    'GraphRSColumn'     : 'NDVI',  # Tree-level Remote Sensing Index Value
    'GraphFieldColumn'  : 'DEPS.1' # Measured Tree Physiolgical Parameter Value
}

print('Analyzing...')
AnayzeObj = ITCAnalyze.AnalyzeUtils(FileDataInfo)
AnayzeObj.ExtractPhenology()
AnayzeObj.plot_spectrum()
AnayzeObj.Generate_RS_REF_Merged_File()
AnayzeObj.ModelPhenologyTreesInMultiImages()
